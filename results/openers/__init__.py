import io
from itertools import islice
from pathlib import Path

import chardet

from ..resultset import Results
from .csv import (  # noqa
    csv_column_names,
    csv_rows_it,
    from_csv,
    sniff_csv_dialect,
    write_csv_to_f,
    write_csv_to_filehandle,
)


def detect_string_enc(contents):
    b = io.BytesIO(contents)
    b.seek(0)
    return detect_enc(b)


def detect_enc(f):
    return chardet.detect(f.read())["encoding"]


def smart_open(f):
    try:
        return io.open(f)
    except TypeError:
        return f


def first_n_lines(stream, n=10):
    return list(islice(stream, n))


def dicts_from_rows(rows):
    try:
        first = rows[0]
    except IndexError:
        return []
    rest = rows[1:]

    def it():
        for d in rest:
            dd = dict(zip(first, d))
            yield dd

    return list(it())


def from_xlsx(f):
    from openpyxl import load_workbook

    if isinstance(f, Path):
        f = str(f)
    wb = load_workbook(filename=f, read_only=True)

    wsheets = list(wb)

    def cell_value(c):
        try:
            internal_value = c.internal_value
            return internal_value or ""
        except AttributeError:
            return ""

    def xget_row_values(row):
        return [cell_value(c) for c in list(row)]

    def do_sheet(ws):
        rows = [xget_row_values(_) for _ in list(ws.rows)]
        return dicts_from_rows(rows)

    return dict(zip(wb.sheetnames, (Results(do_sheet(_)) for _ in wsheets)))


def from_xls(f, file_contents=None):
    from xlrd import open_workbook

    wb = open_workbook(str(f), file_contents=file_contents)

    def get_row_values(ws, rownum):
        return [str(ws.cell_value(rowx=rownum, colx=c) or "") for c in range(ws.ncols)]

    def do_sheet(ws):
        rows = [get_row_values(ws, rx) for rx in range(ws.nrows)]
        return dicts_from_rows(rows)

    wsheets = [wb.sheet_by_index(_) for _ in range(wb.nsheets)]

    return dict(zip(wb.sheet_names(), (Results(do_sheet(_)) for _ in wsheets)))


OPENERS = {
    ".xlsx": from_xlsx,
    ".xls": from_xls,
    ".csv": from_csv,
    ".tsv": from_csv,
    ".psv": from_csv,
}


def from_file(f, **kwargs):
    p = Path(f)

    extension = p.suffix

    try:
        return OPENERS[extension](f, **kwargs)
    except KeyError:
        raise ValueError(f"cannot open a file with extension: {extension}")


def save_xlsx_sheets(sheets_dict, destination):
    from xlsxwriter import Workbook

    workbook = Workbook(destination)

    for k, sheetdata in sheets_dict.items():
        worksheet = workbook.add_worksheet(k)

        rowdata = [sheetdata.keys()] + sheetdata

        for r, row in enumerate(rowdata):
            for c, col in enumerate(row):
                worksheet.write(r, c, col)

    workbook.close()
